import openpyxl
import re
from tkinter import filedialog, messagebox
import customtkinter as ctk
import os
import shutil
from PIL import Image
import datetime

def extract_full_marks(text):
    """Extract full marks from header text like 'Assignment (20)'"""
    if text is None:
        return 0
    match = re.search(r"\((\d+)", str(text))
    return int(match.group(1)) if match else 0

def safe_number(value):
    """Convert cell value to float, return 0 if empty"""
    return float(value) if value not in (None, "") else 0.0

def get_grade(mark):
    """Convert numerical mark to letter grade"""
    if mark >= 90: return "A+"
    elif mark >= 80: return "A"
    elif mark >= 75: return "A-"
    elif mark >= 70: return "B+"
    elif mark >= 60: return "B"
    elif mark >= 60: return "C+"
    elif mark >= 55: return "C"
    elif mark >= 50: return "C-"
    elif mark >= 45: return "D"
    else: return "F"


def process_cqi_data(file_path, assessment_mode):
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # FIXED SHEET LAYOUT
    HEADER_ROW = 6
    DATA_START_ROW = 9

    # AUTO DETECT LAST STUDENT ROW (BASED ON COLUMN E)
    DATA_END_ROW = ws.max_row
    while DATA_END_ROW >= DATA_START_ROW and ws.cell(row=DATA_END_ROW, column=5).value in (None, ""):
        DATA_END_ROW -= 1

    # SUMMARY ROWS (AUTO – ALWAYS BOTTOM)
    TARGET_ROW = DATA_END_ROW + 1
    PERCENT_ROW = DATA_END_ROW + 2

    all_total_marks = []

    # MERGE HEADER CELLS (will be adjusted based on mode)
    if assessment_mode == 3:
        ws.merge_cells("J6:J8")  # Total Marks
        ws.merge_cells("K6:K8")  # Grade
        ws.merge_cells("L6:P6")  # CLO Attainment (5 CLOs)
        ws.merge_cells("Q6:U6")  # PLO Attainment (5 PLOs)
    elif assessment_mode == 2:
        # 4 assessments: Total Marks in col 9, Grade in col 10, CLO starts at col 11
        ws.merge_cells("I6:I8")  # Total Marks
        ws.merge_cells("J6:J8")  # Grade
        ws.merge_cells("K6:N6")  # CLO Attainment (4 CLOs)
        ws.merge_cells("O6:R6")  # PLO Attainment (4 PLOs)
    else:
        # 3 assessments: Total Marks in col 8, Grade in col 9, CLO starts at col 10
        ws.merge_cells("H6:H8")  # Total Marks
        ws.merge_cells("I6:I8")  # Grade
        ws.merge_cells("J6:L6")  # CLO Attainment (3 CLOs)
        ws.merge_cells("M6:O6")  # PLO Attainment (3 PLOs)
    # Merge summary rows based on mode
    if assessment_mode == 3:
        ws.merge_cells(start_row=TARGET_ROW, start_column=1, end_row=TARGET_ROW, end_column=11)
        ws.merge_cells(start_row=PERCENT_ROW, start_column=1, end_row=PERCENT_ROW, end_column=11)
    elif assessment_mode == 2:
        ws.merge_cells(start_row=TARGET_ROW, start_column=1, end_row=TARGET_ROW, end_column=10)
        ws.merge_cells(start_row=PERCENT_ROW, start_column=1, end_row=PERCENT_ROW, end_column=10)
    else:
        ws.merge_cells(start_row=TARGET_ROW, start_column=1, end_row=TARGET_ROW, end_column=9)
        ws.merge_cells(start_row=PERCENT_ROW, start_column=1, end_row=PERCENT_ROW, end_column=9)

    # SET HEADER LABELS (will be adjusted based on mode)
    if assessment_mode == 3:
        ws["J6"] = "Total Marks"
        ws["K6"] = "Grade"
        ws["L6"] = "CLO Attainment"
        ws["Q6"] = "PLO Attainment"
    elif assessment_mode == 2:
        ws["I6"] = "Total Marks"
        ws["J6"] = "Grade"
        ws["K6"] = "CLO Attainment"
        ws["O6"] = "PLO Attainment"
    else:
        ws["H6"] = "Total Marks"
        ws["I6"] = "Grade"
        ws["J6"] = "CLO Attainment"
        ws["M6"] = "PLO Attainment"
    ws.cell(row=TARGET_ROW, column=1).value = "Total Student Achieved Target (>=50%)"
    ws.cell(row=PERCENT_ROW, column=1).value = "% Student Achieved Target"

    # READ MARK HEADERS
    assignment_header = ws.cell(row=HEADER_ROW, column=5).value
    midterm_header = ws.cell(row=HEADER_ROW, column=7).value if assessment_mode == 1 else ws.cell(row=HEADER_ROW, column=7).value
    
    FULL_ASSIGNMENT = extract_full_marks(assignment_header)
    
    if assessment_mode == 3:
        assignment2_header = ws.cell(row=HEADER_ROW, column=6).value
        assignment3_header = ws.cell(row=HEADER_ROW, column=7).value
        midterm_header = ws.cell(row=HEADER_ROW, column=8).value
        final_header = ws.cell(row=HEADER_ROW, column=9).value
        
        FULL_ASSIGNMENT2 = extract_full_marks(assignment2_header)
        FULL_ASSIGNMENT3 = extract_full_marks(assignment3_header)
        FULL_MIDTERM = extract_full_marks(midterm_header)
        FULL_FINAL = extract_full_marks(final_header)
        
        print(f"\n{'='*50}")
        print("FOUR ASSESSMENTS + FINAL MODE")
        print(f"{'='*50}")
        print(f"Assignment 1 Full Marks: {FULL_ASSIGNMENT}")
        print(f"Assignment 2 Full Marks: {FULL_ASSIGNMENT2}")
        print(f"Assignment 3 Full Marks: {FULL_ASSIGNMENT3}")
        print(f"Midterm Full Marks: {FULL_MIDTERM}")
        print(f"Final Exam Full Marks: {FULL_FINAL}")
        print(f"{'='*50}")
        
        # Weight distribution for 5 assessments
        WEIGHT_ASSIGNMENT = 10
        WEIGHT_ASSIGNMENT2 = 10
        WEIGHT_ASSIGNMENT3 = 10
        WEIGHT_MIDTERM = 30
        WEIGHT_FINAL = 40
        
        # ROW 7 - CLO/PLO Labels (starting from column 12)
        # Auto-detect from columns 5 to 9 (E7 to I7)
        clos = []
        plos = []
        for c in range(5, 10):
            val = str(ws.cell(row=7, column=c).value or "").strip()
            
            clo_match = re.search(r'CLO[\s_]*(\d+)', val, re.IGNORECASE)
            plo_match = re.search(r'PLO[\s_]*(\d+)', val, re.IGNORECASE)
            
            if clo_match:
                clos.append(f"CLO{clo_match.group(1)}")
            elif "-" in val:
                clos.append(val.split("-")[0].strip())
            elif val and "PLO" not in val.upper():
                clos.append(val)
            else:
                clos.append(f"CLO{c-4}")
                
            if plo_match:
                plos.append(f"PLO{plo_match.group(1)}")
            elif "-" in val and len(val.split("-")) > 1:
                plos.append(val.split("-")[1].strip())
            else:
                plos.append(f"PLO{c-4}")
        labels = clos + plos
        for i, v in enumerate(labels):
            ws.cell(row=7, column=12 + i, value=v)
        
        # ROW 8 - Assessment percentages (starting from column 12)
        values = [
            f"{FULL_ASSIGNMENT}%", f"{FULL_ASSIGNMENT2}%", f"{FULL_ASSIGNMENT3}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%",
            f"{FULL_ASSIGNMENT}%", f"{FULL_ASSIGNMENT2}%", f"{FULL_ASSIGNMENT3}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%"
        ]
        for i, v in enumerate(values):
            ws.cell(row=8, column=12 + i, value=v)
        
        # Counters for CLO/PLO attainment
        clo1_pass = clo2_pass = clo3_pass = clo4_pass = clo5_pass = 0
        plo1_pass = plo2_pass = plo3_pass = plo4_pass = plo5_pass = 0
        
        failed_a1 = []
        failed_a2 = []
        failed_a3 = []
        failed_m = []
        failed_f = []

        total_students = DATA_END_ROW - DATA_START_ROW + 1
        
        # Process each student
        for r in range(DATA_START_ROW, DATA_END_ROW + 1):
            student_id = str(ws.cell(row=r, column=2).value or "").strip()
            student_name = str(ws.cell(row=r, column=3).value or "").strip()
            if student_name and student_id:
                student_label = f"{student_name} ({student_id})"
            else:
                student_label = student_name or student_id or f"Row {r}"

            a1 = safe_number(ws.cell(row=r, column=5).value)
            a2 = safe_number(ws.cell(row=r, column=6).value)
            a3 = safe_number(ws.cell(row=r, column=7).value)
            m = safe_number(ws.cell(row=r, column=8).value)
            f = safe_number(ws.cell(row=r, column=9).value)
            
            # Calculate CLO percentages (used for individual pass/fail)
            clo1 = (a1 / FULL_ASSIGNMENT) * 100 if FULL_ASSIGNMENT else 0
            clo2 = (a2 / FULL_ASSIGNMENT2) * 100 if FULL_ASSIGNMENT2 else 0
            clo3 = (a3 / FULL_ASSIGNMENT3) * 100 if FULL_ASSIGNMENT3 else 0
            clo4 = (m / FULL_MIDTERM) * 100 if FULL_MIDTERM else 0
            clo5 = (f / FULL_FINAL) * 100 if FULL_FINAL else 0
            
            if clo1 < 50: failed_a1.append(student_label)
            if clo2 < 50: failed_a2.append(student_label)
            if clo3 < 50: failed_a3.append(student_label)
            if clo4 < 50: failed_m.append(student_label)
            if clo5 < 50: failed_f.append(student_label)

            # Calculate total mark with new weights
            total_mark = (a1 / FULL_ASSIGNMENT * WEIGHT_ASSIGNMENT if FULL_ASSIGNMENT else 0) + \
                        (a2 / FULL_ASSIGNMENT2 * WEIGHT_ASSIGNMENT2 if FULL_ASSIGNMENT2 else 0) + \
                        (a3 / FULL_ASSIGNMENT3 * WEIGHT_ASSIGNMENT3 if FULL_ASSIGNMENT3 else 0) + \
                        (m / FULL_MIDTERM * WEIGHT_MIDTERM if FULL_MIDTERM else 0) + \
                        (f / FULL_FINAL * WEIGHT_FINAL if FULL_FINAL else 0)
            total_mark = round(total_mark, 2)
            all_total_marks.append(total_mark)
            
            # Save PASS/FAIL status in column 4 (Column D) - Fails if any individual assessment fails or total mark < 50
            if total_mark >= 50 and clo1 >= 50 and clo2 >= 50 and clo3 >= 50 and clo4 >= 50 and clo5 >= 50:
                ws.cell(row=r, column=4).value = "PASS"
            else:
                ws.cell(row=r, column=4).value = "FAIL"
            
            ws.cell(row=r, column=10).value = total_mark
            ws.cell(row=r, column=11).value = get_grade(total_mark)
            
            # Write CLO and PLO values (columns 12-21)
            ws.cell(row=r, column=12).value = round(clo1, 2)
            ws.cell(row=r, column=13).value = round(clo2, 2)
            ws.cell(row=r, column=14).value = round(clo3, 2)
            ws.cell(row=r, column=15).value = round(clo4, 2)
            ws.cell(row=r, column=16).value = round(clo5, 2)
            ws.cell(row=r, column=17).value = round(clo1, 2)
            ws.cell(row=r, column=18).value = round(clo2, 2)
            ws.cell(row=r, column=19).value = round(clo3, 2)
            ws.cell(row=r, column=20).value = round(clo4, 2)
            ws.cell(row=r, column=21).value = round(clo5, 2)
            
            # Count students achieving target (>=50%)
            if clo1 >= 50: clo1_pass += 1
            if clo2 >= 50: clo2_pass += 1
            if clo3 >= 50: clo3_pass += 1
            if clo4 >= 50: clo4_pass += 1
            if clo5 >= 50: clo5_pass += 1
            if clo1 >= 50: plo1_pass += 1
            if clo2 >= 50: plo2_pass += 1
            if clo3 >= 50: plo3_pass += 1
            if clo4 >= 50: plo4_pass += 1
            if clo5 >= 50: plo5_pass += 1

    elif assessment_mode == 2:
        # Three assessments + Final: Assignment, Assignment2, Midterm, Final
        assignment2_header = ws.cell(row=HEADER_ROW, column=6).value
        midterm_header = ws.cell(row=HEADER_ROW, column=7).value
        final_header = ws.cell(row=HEADER_ROW, column=8).value
        
        FULL_ASSIGNMENT2 = extract_full_marks(assignment2_header)
        FULL_MIDTERM = extract_full_marks(midterm_header)
        FULL_FINAL = extract_full_marks(final_header)
        
        print(f"\n{'='*50}")
        print("THREE ASSESSMENTS + FINAL MODE")
        print(f"{'='*50}")
        print(f"Assignment 1 Full Marks: {FULL_ASSIGNMENT}")
        print(f"Assignment 2 Full Marks: {FULL_ASSIGNMENT2}")
        print(f"Midterm Full Marks: {FULL_MIDTERM}")
        print(f"Final Exam Full Marks: {FULL_FINAL}")
        print(f"{'='*50}")
        
        # Weight distribution for 4 assessments
        WEIGHT_ASSIGNMENT = 15
        WEIGHT_ASSIGNMENT2 = 15
        WEIGHT_MIDTERM = 30
        WEIGHT_FINAL = 40
        
        # ROW 7 - CLO/PLO Labels (starting from column 11)
        # Auto-detect from columns 5 to 8 (E7 to H7)
        clos = []
        plos = []
        for c in range(5, 9):
            val = str(ws.cell(row=7, column=c).value or "").strip()
            
            clo_match = re.search(r'CLO[\s_]*(\d+)', val, re.IGNORECASE)
            plo_match = re.search(r'PLO[\s_]*(\d+)', val, re.IGNORECASE)
            
            if clo_match:
                clos.append(f"CLO{clo_match.group(1)}")
            elif "-" in val:
                clos.append(val.split("-")[0].strip())
            elif val and "PLO" not in val.upper():
                clos.append(val)
            else:
                clos.append(f"CLO{c-4}")
                
            if plo_match:
                plos.append(f"PLO{plo_match.group(1)}")
            elif "-" in val and len(val.split("-")) > 1:
                plos.append(val.split("-")[1].strip())
            else:
                plos.append(f"PLO{c-4}")
        labels = clos + plos
        for i, v in enumerate(labels):
            ws.cell(row=7, column=11 + i, value=v)
        
        # ROW 8 - Assessment percentages (starting from column 11)
        values = [
            f"{FULL_ASSIGNMENT}%", f"{FULL_ASSIGNMENT2}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%",
            f"{FULL_ASSIGNMENT}%", f"{FULL_ASSIGNMENT2}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%"
        ]
        for i, v in enumerate(values):
            ws.cell(row=8, column=11 + i, value=v)
        
        # Counters for CLO/PLO attainment
        clo1_pass = clo2_pass = clo3_pass = clo4_pass = 0
        plo1_pass = plo2_pass = plo3_pass = plo6_pass = 0
        
        failed_a1 = []
        failed_a2 = []
        failed_m = []
        failed_f = []
        
        total_students = DATA_END_ROW - DATA_START_ROW + 1
        
        # Process each student
        for r in range(DATA_START_ROW, DATA_END_ROW + 1):
            student_id = str(ws.cell(row=r, column=2).value or "").strip()
            student_name = str(ws.cell(row=r, column=3).value or "").strip()
            if student_name and student_id:
                student_label = f"{student_name} ({student_id})"
            else:
                student_label = student_name or student_id or f"Row {r}"

            a1 = safe_number(ws.cell(row=r, column=5).value)
            a2 = safe_number(ws.cell(row=r, column=6).value)
            m = safe_number(ws.cell(row=r, column=7).value)
            f = safe_number(ws.cell(row=r, column=8).value)
            
            # Calculate CLO percentages (used for individual pass/fail)
            clo1 = (a1 / FULL_ASSIGNMENT) * 100
            if clo1 < 50: failed_a1.append(student_label)
            
            clo2 = (a2 / FULL_ASSIGNMENT2) * 100
            if clo2 < 50: failed_a2.append(student_label)
            
            clo3 = (m / FULL_MIDTERM) * 100
            if clo3 < 50: failed_m.append(student_label)
            
            clo4 = (f / FULL_FINAL) * 100
            if clo4 < 50: failed_f.append(student_label)
            
            # Calculate total mark with new weights
            total_mark = ((a1 / FULL_ASSIGNMENT) * WEIGHT_ASSIGNMENT) + \
                        ((a2 / FULL_ASSIGNMENT2) * WEIGHT_ASSIGNMENT2) + \
                        ((m / FULL_MIDTERM) * WEIGHT_MIDTERM) + \
                        ((f / FULL_FINAL) * WEIGHT_FINAL)
            total_mark = round(total_mark, 2)
            all_total_marks.append(total_mark)
            
            # Save PASS/FAIL status in column 4 (Column D) - Fails if any individual assessment fails or total mark < 50
            if total_mark >= 50 and clo1 >= 50 and clo2 >= 50 and clo3 >= 50 and clo4 >= 50:
                ws.cell(row=r, column=4).value = "PASS"
            else:
                ws.cell(row=r, column=4).value = "FAIL"
            
            ws.cell(row=r, column=9).value = total_mark
            ws.cell(row=r, column=10).value = get_grade(total_mark)
            
            # Write CLO and PLO values (columns 11-18)
            ws.cell(row=r, column=11).value = round(clo1, 2)  # CLO1
            ws.cell(row=r, column=12).value = round(clo2, 2)  # CLO2
            ws.cell(row=r, column=13).value = round(clo3, 2)  # CLO3
            ws.cell(row=r, column=14).value = round(clo4, 2)  # CLO4
            ws.cell(row=r, column=15).value = round(clo1, 2)  # PLO1
            ws.cell(row=r, column=16).value = round(clo2, 2)  # PLO2
            ws.cell(row=r, column=17).value = round(clo3, 2)  # PLO3
            ws.cell(row=r, column=18).value = round(clo4, 2)  # PLO6
            
            # Count students achieving target (>=50%)
            if clo1 >= 50: clo1_pass += 1
            if clo2 >= 50: clo2_pass += 1
            if clo3 >= 50: clo3_pass += 1
            if clo4 >= 50: clo4_pass += 1
            if clo1 >= 50: plo1_pass += 1
            if clo2 >= 50: plo2_pass += 1
            if clo3 >= 50: plo3_pass += 1
            if clo4 >= 50: plo6_pass += 1

        
    else:  # assessment_mode == 1
        # Two assessments + Final: Assignment, Midterm, Final
        midterm_header = ws.cell(row=HEADER_ROW, column=6).value
        final_header = ws.cell(row=HEADER_ROW, column=7).value
        
        FULL_MIDTERM = extract_full_marks(midterm_header)
        FULL_FINAL = extract_full_marks(final_header)
        
        print(f"\n{'='*50}")
        print("TWO ASSESSMENTS + FINAL MODE")
        print(f"{'='*50}")
        print(f"Assignment Full Marks: {FULL_ASSIGNMENT}")
        print(f"Midterm Full Marks: {FULL_MIDTERM}")
        print(f"Final Exam Full Marks: {FULL_FINAL}")
        print(f"{'='*50}")
        
        # Weight distribution for 3 assessments (matching original code)
        WEIGHT_ASSIGNMENT = 20
        WEIGHT_MIDTERM = 30
        WEIGHT_FINAL = 50
        
        # ROW 7 - CLO/PLO Labels
        # Auto-detect from columns 5 to 7 (E7 to G7)
        clos = []
        plos = []
        for c in range(5, 8):
            val = str(ws.cell(row=7, column=c).value or "").strip()
            
            clo_match = re.search(r'CLO[\s_]*(\d+)', val, re.IGNORECASE)
            plo_match = re.search(r'PLO[\s_]*(\d+)', val, re.IGNORECASE)
            
            if clo_match:
                clos.append(f"CLO{clo_match.group(1)}")
            elif "-" in val:
                clos.append(val.split("-")[0].strip())
            elif val and "PLO" not in val.upper():
                clos.append(val)
            else:
                clos.append(f"CLO{c-4}")
                
            if plo_match:
                plos.append(f"PLO{plo_match.group(1)}")
            elif "-" in val and len(val.split("-")) > 1:
                plos.append(val.split("-")[1].strip())
            else:
                plos.append(f"PLO{c-4}")
        labels = clos + plos
        for i, v in enumerate(labels):
            ws.cell(row=7, column=10 + i, value=v)
        
        # ROW 8 - Assessment percentages
        values = [
            f"{FULL_ASSIGNMENT}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%",
            f"{FULL_ASSIGNMENT}%", f"{FULL_MIDTERM}%", f"{FULL_FINAL}%"
        ]
        for i, v in enumerate(values):
            ws.cell(row=8, column=10 + i, value=v)
        
        # Counters for CLO/PLO attainment
        clo1_pass = clo2_pass = clo3_pass = 0
        plo1_pass = plo2_pass = plo6_pass = 0
        
        failed_a = []
        failed_m = []
        failed_f = []
        
        total_students = DATA_END_ROW - DATA_START_ROW + 1
        
        # Process each student
        for r in range(DATA_START_ROW, DATA_END_ROW + 1):
            student_id = str(ws.cell(row=r, column=2).value or "").strip()
            student_name = str(ws.cell(row=r, column=3).value or "").strip()
            if student_name and student_id:
                student_label = f"{student_name} ({student_id})"
            else:
                student_label = student_name or student_id or f"Row {r}"

            a = safe_number(ws.cell(row=r, column=5).value)
            m = safe_number(ws.cell(row=r, column=6).value)
            f = safe_number(ws.cell(row=r, column=7).value)
            
            # Calculate CLO percentages (used for individual pass/fail)
            clo1 = (a / FULL_ASSIGNMENT) * 100
            if clo1 < 50: failed_a.append(student_label)
            
            clo2 = (m / FULL_MIDTERM) * 100
            if clo2 < 50: failed_m.append(student_label)
            
            clo3 = (f / FULL_FINAL) * 100
            if clo3 < 50: failed_f.append(student_label)
            
            # Calculate total mark with weights (matching original code)
            total_mark = ((a / FULL_ASSIGNMENT) * WEIGHT_ASSIGNMENT) + \
                        ((m / FULL_MIDTERM) * WEIGHT_MIDTERM) + \
                        ((f / FULL_FINAL) * WEIGHT_FINAL)
            total_mark = round(total_mark, 2)
            all_total_marks.append(total_mark)
            
            # Save PASS/FAIL status in column 4 (Column D) - Fails if any individual assessment fails or total mark < 50
            if total_mark >= 50 and clo1 >= 50 and clo2 >= 50 and clo3 >= 50:
                ws.cell(row=r, column=4).value = "PASS"
            else:
                ws.cell(row=r, column=4).value = "FAIL"
            
            ws.cell(row=r, column=8).value = total_mark
            ws.cell(row=r, column=9).value = get_grade(total_mark)
            
            # Write CLO and PLO values
            ws.cell(row=r, column=10).value = round(clo1, 2)
            ws.cell(row=r, column=11).value = round(clo2, 2)
            ws.cell(row=r, column=12).value = round(clo3, 2)
            ws.cell(row=r, column=13).value = round(clo1, 2)  # PLO1 = CLO1
            ws.cell(row=r, column=14).value = round(clo2, 2)  # PLO2 = CLO2
            ws.cell(row=r, column=15).value = round(clo3, 2)  # PLO6 = CLO3
            
            # Count students achieving target (>=50%)
            if clo1 >= 50: clo1_pass += 1
            if clo2 >= 50: clo2_pass += 1
            if clo3 >= 50: clo3_pass += 1
            if clo1 >= 50: plo1_pass += 1
            if clo2 >= 50: plo2_pass += 1
            if clo3 >= 50: plo6_pass += 1

    
    # Write summary rows
    if assessment_mode == 3:
        # 5 assessments mode summary (columns 12-21)
        ws.cell(row=TARGET_ROW, column=12).value = clo1_pass
        ws.cell(row=TARGET_ROW, column=13).value = clo2_pass
        ws.cell(row=TARGET_ROW, column=14).value = clo3_pass
        ws.cell(row=TARGET_ROW, column=15).value = clo4_pass
        ws.cell(row=TARGET_ROW, column=16).value = clo5_pass
        ws.cell(row=TARGET_ROW, column=17).value = plo1_pass
        ws.cell(row=TARGET_ROW, column=18).value = plo2_pass
        ws.cell(row=TARGET_ROW, column=19).value = plo3_pass
        ws.cell(row=TARGET_ROW, column=20).value = plo4_pass
        ws.cell(row=TARGET_ROW, column=21).value = plo5_pass
        
        ws.cell(row=PERCENT_ROW, column=12).value = f"{int((clo1_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=13).value = f"{int((clo2_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=14).value = f"{int((clo3_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=15).value = f"{int((clo4_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=16).value = f"{int((clo5_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=17).value = f"{int((plo1_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=18).value = f"{int((plo2_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=19).value = f"{int((plo3_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=20).value = f"{int((plo4_pass/total_students)*100)}%" if total_students else "0%"
        ws.cell(row=PERCENT_ROW, column=21).value = f"{int((plo5_pass/total_students)*100)}%" if total_students else "0%"
    elif assessment_mode == 2:
        # 4 assessments mode summary (columns 11-18)
        ws.cell(row=TARGET_ROW, column=11).value = clo1_pass
        ws.cell(row=TARGET_ROW, column=12).value = clo2_pass
        ws.cell(row=TARGET_ROW, column=13).value = clo3_pass
        ws.cell(row=TARGET_ROW, column=14).value = clo4_pass
        ws.cell(row=TARGET_ROW, column=15).value = plo1_pass
        ws.cell(row=TARGET_ROW, column=16).value = plo2_pass
        ws.cell(row=TARGET_ROW, column=17).value = plo3_pass
        ws.cell(row=TARGET_ROW, column=18).value = plo6_pass
        
        ws.cell(row=PERCENT_ROW, column=11).value = f"{int((clo1_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=12).value = f"{int((clo2_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=13).value = f"{int((clo3_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=14).value = f"{int((clo4_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=15).value = f"{int((plo1_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=16).value = f"{int((plo2_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=17).value = f"{int((plo3_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=18).value = f"{int((plo6_pass/total_students)*100)}%"
    else:
        # 3 assessments mode summary
        ws.cell(row=TARGET_ROW, column=10).value = clo1_pass
        ws.cell(row=TARGET_ROW, column=11).value = clo2_pass
        ws.cell(row=TARGET_ROW, column=12).value = clo3_pass
        ws.cell(row=TARGET_ROW, column=13).value = plo1_pass
        ws.cell(row=TARGET_ROW, column=14).value = plo2_pass
        ws.cell(row=TARGET_ROW, column=15).value = plo6_pass
        
        ws.cell(row=PERCENT_ROW, column=10).value = f"{int((clo1_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=11).value = f"{int((clo2_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=12).value = f"{int((clo3_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=13).value = f"{int((plo1_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=14).value = f"{int((plo2_pass/total_students)*100)}%"
        ws.cell(row=PERCENT_ROW, column=15).value = f"{int((plo6_pass/total_students)*100)}%"

    
    FAIL_SUMMARY_START_ROW = PERCENT_ROW + 2
    ws.cell(row=FAIL_SUMMARY_START_ROW, column=1).value = "SUMMARY OF FAILED STUDENTS (<50%)"
    try:
        from openpyxl.styles import Font
        ws.cell(row=FAIL_SUMMARY_START_ROW, column=1).font = Font(bold=True)
    except:
        pass
        
    current_row = FAIL_SUMMARY_START_ROW + 2
    
    def write_failed_list(title, failed_list, start_row):
        ws.cell(row=start_row, column=1).value = title
        try:
            ws.cell(row=start_row, column=1).font = Font(bold=True)
        except:
            pass
        row = start_row + 1
        if not failed_list:
            ws.cell(row=row, column=1).value = "None"
            row += 1
        else:
            for s in failed_list:
                ws.cell(row=row, column=1).value = s
                row += 1
        return row + 1

    if assessment_mode == 3:
        current_row = write_failed_list("Assignment 1 Failures:", failed_a1, current_row)
        current_row = write_failed_list("Assignment 2 Failures:", failed_a2, current_row)
        current_row = write_failed_list("Assignment 3 Failures:", failed_a3, current_row)
        current_row = write_failed_list("Midterm Failures:", failed_m, current_row)
        current_row = write_failed_list("Final Exam Failures:", failed_f, current_row)
    elif assessment_mode == 2:
        current_row = write_failed_list("Assignment 1 Failures:", failed_a1, current_row)
        current_row = write_failed_list("Assignment 2 Failures:", failed_a2, current_row)
        current_row = write_failed_list("Midterm Failures:", failed_m, current_row)
        current_row = write_failed_list("Final Exam Failures:", failed_f, current_row)
    else:
        current_row = write_failed_list("Assignment Failures:", failed_a, current_row)
        current_row = write_failed_list("Midterm Failures:", failed_m, current_row)
        current_row = write_failed_list("Final Exam Failures:", failed_f, current_row)

    # ADD MARKS SUMMARY TO COLUMN 4
    ws.cell(row=FAIL_SUMMARY_START_ROW, column=4).value = "Marks Summary:"
    try:
        from openpyxl.styles import Font
        ws.cell(row=FAIL_SUMMARY_START_ROW, column=4).font = Font(bold=True)
    except:
        pass
        
    if all_total_marks:
        avg_marks = round(sum(all_total_marks) / len(all_total_marks), 2)
        max_marks = max(all_total_marks)
        min_marks = min(all_total_marks)
    else:
        avg_marks = max_marks = min_marks = 0
        
    ws.cell(row=FAIL_SUMMARY_START_ROW + 1, column=4).value = f"Average marks: {avg_marks}"
    ws.cell(row=FAIL_SUMMARY_START_ROW + 2, column=4).value = f"Highest marks: {max_marks}"
    ws.cell(row=FAIL_SUMMARY_START_ROW + 3, column=4).value = f"Lowest marks: {min_marks}"

    # ADD GRADE SUMMARY TO COLUMN 7
    ws.cell(row=FAIL_SUMMARY_START_ROW, column=7).value = "Number of students per grade:"
    try:
        ws.cell(row=FAIL_SUMMARY_START_ROW, column=7).font = Font(bold=True)
    except:
        pass
        
    if all_total_marks:
        all_grades = [get_grade(m) for m in all_total_marks]
        grade_order = ['A+', 'A', 'A-', 'B+', 'B', 'C+', 'C', 'C-', 'D', 'F']
        grade_counts = {g: all_grades.count(g) for g in grade_order}
        
        row_offset = 1
        for g in grade_order:
            ws.cell(row=FAIL_SUMMARY_START_ROW + row_offset, column=7).value = f"{g}: {grade_counts[g]}"
            row_offset += 1

    # Save the file
    base_name, ext = os.path.splitext(file_path)
    output_file = f"{base_name}_updated{ext}"
    wb.save(output_file)
    
    print(f"\n{'='*50}")
    print(f"> Processing complete!")
    print(f"> Total students processed: {total_students}")
    print(f"> Output saved to: {output_file}")
    print(f"{'='*50}\n")
    return total_students, output_file

class CQIApp(ctk.CTk):

    def __init__(self):
        super().__init__()

        # Window settings
        self.title("UCSI Automated Student Assessment System")
        self.geometry("720x760")
        self.minsize(650, 720)

        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.selected_file = None
        self.history = []

        # Main container
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=25, pady=25)

        # ================= THEME SWITCH (Top Left) =================
        self.theme_switch = ctk.CTkSwitch(
            self,
            text="Dark Mode",
            command=self.toggle_theme
        )
        self.theme_switch.place(x=20, y=20)
        
        # ================= HELP BUTTON (Top Right) =================
        self.help_btn = ctk.CTkButton(
            self,
            text="ℹ️ Help",
            width=70,
            height=28,
            corner_radius=15,
            fg_color="transparent",
            border_width=1,
            text_color=("gray10", "gray90"),
            command=self.show_help_message
        )
        self.help_btn.place(relx=1.0, x=-20, y=20, anchor="ne")

        # ================= HEADER =================
        self.header_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.header_frame.pack(fill="x", pady=(0,10))

        try:
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ucsilogo.png")
            if os.path.exists(logo_path):

                pil_image = Image.open(logo_path)

                width = 220
                ratio = width / pil_image.size[0]
                height = int(pil_image.size[1] * ratio)

                self.logo_image = ctk.CTkImage(
                    light_image=pil_image,
                    dark_image=pil_image,
                    size=(width, height)
                )

                self.logo_label = ctk.CTkLabel(
                    self.header_frame,
                    image=self.logo_image,
                    text=""
                )
                self.logo_label.pack(pady=(0,5))

        except:
            pass

        self.title1 = ctk.CTkLabel(
            self.header_frame,
            text="UCSI University",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=("gray40","gray60")
        )
        self.title1.pack()

        self.title2 = ctk.CTkLabel(
            self.header_frame,
            text="Automated Student Assessment System",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        self.title2.pack(pady=(5,5))

        # ================= MODE CARD =================
        self.mode_card = ctk.CTkFrame(self.main_frame, corner_radius=18)
        self.mode_card.pack(fill="x", padx=60, pady=5)

        self.mode_header_frame = ctk.CTkFrame(self.mode_card, fg_color="transparent")
        self.mode_header_frame.pack(pady=(10, 10))

        self.mode_label = ctk.CTkLabel(
            self.mode_header_frame,
            text="Select Assessment Mode",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.mode_label.pack(side="left", padx=(0, 5))

        self.mode_info_btn = ctk.CTkButton(
            self.mode_header_frame,
            text="❓",
            width=28,
            height=28,
            corner_radius=14,
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray85", "gray30"),
            command=self.show_mode_info
        )
        self.mode_info_btn.pack(side="left")

        self.mode_var = ctk.IntVar(value=1)

        self.radio1 = ctk.CTkRadioButton(
            self.mode_card,
            text="📊 Two Assessments + Final",
            variable=self.mode_var,
            value=1,
            font=ctk.CTkFont(size=15)
        )
        self.radio1.pack(anchor="w", padx=40, pady=5)

        self.radio2 = ctk.CTkRadioButton(
            self.mode_card,
            text="📈 Three Assessments + Final",
            variable=self.mode_var,
            value=2,
            font=ctk.CTkFont(size=15)
        )
        self.radio2.pack(anchor="w", padx=40, pady=5)

        self.radio3 = ctk.CTkRadioButton(
            self.mode_card,
            text="📑 Four Assessments + Final",
            variable=self.mode_var,
            value=3,
            font=ctk.CTkFont(size=15)
        )
        self.radio3.pack(anchor="w", padx=40, pady=(5,10))

        # ================= FILE SECTION =================
        self.file_frame = ctk.CTkFrame(self.main_frame, corner_radius=15)
        self.file_frame.pack(fill="x", padx=60, pady=5)

        self.file_btn = ctk.CTkButton(
            self.file_frame,
            text="📂 Select Excel File",
            height=45,
            corner_radius=12,
            command=self.select_file
        )
        self.file_btn.pack(pady=(10,5))

        self.file_label = ctk.CTkLabel(
            self.file_frame,
            text="No file selected",
            text_color="gray"
        )
        self.file_label.pack(pady=(0,10))

        # ================= BUTTON SECTION =================
        self.button_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.button_frame.pack(fill="x", padx=60, pady=5)

        self.process_btn = ctk.CTkButton(
            self.button_frame,
            text="🚀 Process Assessment Data",
            height=50,
            corner_radius=12,
            font=ctk.CTkFont(size=16, weight="bold"),
            command=self.process_assessment
        )
        self.process_btn.pack(fill="x", pady=(0,10))

        self.sample_btn = ctk.CTkButton(
            self.button_frame,
            text="📥 Download Sample Template",
            height=50,
            corner_radius=12,
            fg_color="transparent",
            border_width=2,
            text_color=("gray10", "#DCE4EE"),
            font=ctk.CTkFont(size=15),
            command=self.download_sample
        )
        self.sample_btn.pack(fill="x")

        # ================= PROGRESS =================
        self.progress = ctk.CTkProgressBar(self.main_frame, height=6)
        self.progress.pack(fill="x", padx=60, pady=4)
        self.progress.set(0)

        # ================= STATUS =================
        self.status_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.status_frame.pack(fill="x", padx=60, pady=(0, 5))

        self.history_btn = ctk.CTkButton(
            self.status_frame,
            text="🕒 History",
            width=70,
            height=28,
            corner_radius=12,
            fg_color="transparent",
            border_width=1,
            text_color=("gray10", "gray90"),
            command=self.show_history
        )
        self.history_btn.pack(side="right")

        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="Status: Ready",
            text_color="gray"
        )
        self.status_label.pack(side="left", expand=True, padx=(70, 0))

        # ================= FOOTER =================
        self.footer = ctk.CTkLabel(
            self.main_frame,
            text="© UCSI University Assessment Automation System",
            font=ctk.CTkFont(size=11),
            text_color="gray"
        )
        self.footer.pack(pady=(10,0))


    # ================= FILE SELECT =================
    def select_file(self):

        file_path = filedialog.askopenfilename(
            title="Select CQI Assessment Excel File",
            filetypes=[("Excel files","*.xlsx *.xls")]
        )

        if file_path:
            self.selected_file = file_path
            self.file_label.configure(text=os.path.basename(file_path))


    # ================= PROCESS =================
    def process_assessment(self):

        if not self.selected_file:
            messagebox.showwarning("Warning","Please select an Excel file first")
            return

        try:

            self.status_label.configure(text="Processing...", text_color="blue")
            self.progress.set(0.2)
            self.update()

            total_students, output_file = process_cqi_data(
                self.selected_file,
                self.mode_var.get()
            )

            self.progress.set(1)

            self.status_label.configure(
                text=f"✅ Success! {os.path.basename(output_file)} created",
                text_color="green"
            )

            time_str = datetime.datetime.now().strftime("%I:%M %p")
            self.history.insert(0, f"[{time_str}] ✅ Processed: {os.path.basename(output_file)} ({total_students} students)")

            messagebox.showinfo(
                "Success",
                f"Processing complete!\n\nStudents processed: {total_students}\n\nSaved to:\n{output_file}"
            )

        except Exception as e:
            self.progress.set(0)
            
            time_str = datetime.datetime.now().strftime("%I:%M %p")
            self.history.insert(0, f"[{time_str}] ❌ Error Processing: {str(e)}")

            messagebox.showerror("Error", str(e))
            self.status_label.configure(
                text="❌ Error occurred",
                text_color="red"
            )

    def download_sample(self):
        mode = self.mode_var.get()
        if mode == 1:
            source = "CQI_SAMPLE_FORMAT.xlsx"
        elif mode == 2:
            source = "CQI_SAMPLE_FORMAT 2.xlsx"
        else:
            source = "CQI_SAMPLE_FORMAT 3.xlsx"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx")]
        )
        if not save_path:
            return

        try:
            shutil.copyfile(source, save_path)
            self.status_label.configure(
                text="Sample template downloaded",
                text_color="green"
            )
            time_str = datetime.datetime.now().strftime("%I:%M %p")
            self.history.insert(0, f"[{time_str}] 📥 Downloaded Sample: {os.path.basename(save_path)}")
            messagebox.showinfo(
                "Success",
                "Sample template downloaded successfully"
            )
        except Exception as e:
            time_str = datetime.datetime.now().strftime("%I:%M %p")
            self.history.insert(0, f"[{time_str}] ❌ Error Downloading Sample: {str(e)}")
            messagebox.showerror("Error", str(e))

    # ================= INFO / HELP MESSAGES =================
    def show_history(self):
        if not self.history:
            messagebox.showinfo("Processing History", "No history available yet.")
            return
            
        history_text = "\n".join(self.history)
        messagebox.showinfo("Processing History", history_text)

    def show_help_message(self):
        help_text = (
            "Welcome to the Assessment Automation System!\n\n"
            "Step 1: Choose your Assessment Mode based on the number of coursework assessments.\n"
            "Step 2: Download a Sample Template if you need to see the required Excel format.\n"
            "Step 3: Select your filled Excel file containing student marks.\n"
            "Step 4: Click 'Process Assessment Data' to generate results.\n\n"
            "The system calculates totals, assigns grades, determines pass/fail status, "
            "and evaluates CLO/PLO attainment rates automatically."
        )
        messagebox.showinfo("Instructions & Help", help_text)

    def show_mode_info(self):
        info_text = (
            "Assessment Modes:\n\n"
            "📊 Two Assessments + Final:\n"
            "- Designed for courses with 1 Assignment, 1 Midterm, and 1 Final Exam.\n\n"
            "📈 Three Assessments + Final:\n"
            "- Designed for courses with 2 Assignments, 1 Midterm, and 1 Final Exam.\n\n"
            "📑 Four Assessments + Final:\n"
            "- Designed for courses with 3 Assignments, 1 Midterm, and 1 Final Exam.\n\n"
            "Please ensure your uploaded Excel file corresponds to the selected format."
        )
        messagebox.showinfo("Assessment Mode Information", info_text)

    # ================= THEME SWITCH =================
    def toggle_theme(self):
        if self.theme_switch.get():
            ctk.set_appearance_mode("dark")
        else:
            ctk.set_appearance_mode("light")

def main():
    app = CQIApp()
    app.mainloop()

if __name__ == "__main__":
    main()
